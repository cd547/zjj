#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel读取模块，用于读取Excel文件中的数据
"""

import openpyxl


class ExcelReader:
    """Excel文件读取器"""
    
    def __init__(self, excel_file, header_row=16, start_row=17, end_row=None):
        """初始化ExcelReader
        
        Args:
            excel_file (str): Excel文件路径
            header_row (int): 表头所在行
            start_row (int): 数据开始行
            end_row (int): 数据结束行（None表示到文件末尾）
        """
        self.excel_file = excel_file
        self.header_row = header_row
        self.start_row = start_row
        self.end_row = end_row
    
    def read_data(self):
        """读取Excel文件中的数据
        
        Returns:
            tuple: (data, discount_value)，data是包含数据的列表，discount_value是找到的折扣值
        """
        try:
            # 打开Excel文件
            wb = openpyxl.load_workbook(self.excel_file)
            # 获取第一个工作表
            sheet = wb.active
            
            # 获取工作表的总行数
            max_row = sheet.max_row
            
            # 读取列名
            headers = []
            for cell in sheet[self.header_row]:  # 表头所在行
                headers.append(cell.value)
            
            # 验证表头
            expected_headers = [
                "No.",
                "IMPA",
                "Description",
                "Remark",
                "Q'ty",
                "Unit",
                "Unit Price   (USD)",
                "Amount   (USD)",
                "Brand"
            ]
            
            # 清理实际表头中的空格，以便准确比较
            cleaned_headers = []
            for header in headers:
                if header is not None:
                    # 替换多个连续空格为单个空格，并去除首尾空格
                    import re
                    cleaned = re.sub(r'\s+', ' ', str(header)).strip()
                    cleaned_headers.append(cleaned)
                else:
                    cleaned_headers.append(None)
            
            # 清理预期表头中的空格
            cleaned_expected = []
            for header in expected_headers:
                import re
                cleaned = re.sub(r'\s+', ' ', str(header)).strip()
                cleaned_expected.append(cleaned)
            
            # 验证表头是否一致
            header_match = True
            error_message = ""
            
            # 检查表头数量
            if len(cleaned_headers) < len(cleaned_expected):
                header_match = False
                error_message = f"表头数量不足，预期 {len(cleaned_expected)} 列，实际只有 {len(cleaned_headers)} 列"
            else:
                # 检查每列表头
                for i, (actual, expected) in enumerate(zip(cleaned_headers, cleaned_expected)):
                    if actual != expected:
                        header_match = False
                        error_message = f"表头第 {i+1} 列不匹配，预期: '{expected}'，实际: '{actual}'"
                        break
            
            # 如果表头验证失败，抛出异常
            if not header_match:
                raise ValueError(f"Excel表头验证失败: {error_message}")
            
            print("Excel表头验证通过！")
            
            # 读取数据
            data = []
            # 确定结束行
            actual_end_row = self.end_row
            
            # 如果没有指定end_row，则自动计算
            if actual_end_row is None:
                actual_end_row = max_row  # 默认到文件末尾
                
                # 从start_row开始检查，找到第一个第一列为空或不是数字的行
                for row_num in range(self.start_row, max_row + 1):
                    # 获取第一列（A列）的值
                    first_cell_value = sheet.cell(row=row_num, column=1).value
                    
                    # 检查第一列是否为空或不是数字
                    if first_cell_value is None:
                        # 第一列为空，上一行就是end_row
                        actual_end_row = row_num - 1
                        break
                    elif not isinstance(first_cell_value, (int, float)):
                        # 第一列不是数字，上一行就是end_row
                        actual_end_row = row_num - 1
                        break
                    elif isinstance(first_cell_value, str) and not first_cell_value.strip().isdigit():
                        # 第一列是字符串但不是数字，上一行就是end_row
                        actual_end_row = row_num - 1
                        break
            
            # 确保actual_end_row至少为start_row
            if actual_end_row < self.start_row:
                actual_end_row = self.start_row

            for row in sheet.iter_rows(min_row=self.start_row, max_row=actual_end_row, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] is not None:
                        row_data[headers[i]] = value
                # 只添加非空行
                if any(row_data.values()):
                    data.append(row_data)
            
            # 查找discount值
            discount_value = None
            # 从表格的最后一行开始往上查找
            for row_num in range(max_row, 1, -1):
                # 检查F列（第6列）的值是否为"Discount  :"
                f_cell_value = sheet.cell(row=row_num, column=6).value
                if f_cell_value == "Discount  :":
                    # 检查H列（第8列）的值是否为百分号的数字
                    h_cell_value = sheet.cell(row=row_num, column=8).value
                    if h_cell_value:
                        # 检查是否为字符串且包含百分号
                        if isinstance(h_cell_value, str) and "%" in h_cell_value:
                            try:
                                # 提取百分号前的数字部分
                                discount_str = h_cell_value.replace("%", "").strip()
                                # 转换为浮点数
                                discount_float = float(discount_str)
                                # 转换为整数
                                discount_value = int(discount_float)
                                print(f"找到Discount值: {discount_value}")
                                break
                            except:
                                # 如果转换失败，继续查找
                                pass
                        # 检查是否为数字类型（可能已经是数字格式）
                        elif isinstance(h_cell_value, (int, float)):
                            # 检查是否是小于1的小数（可能是百分比的数字表示，如0.16表示16%）
                            if h_cell_value < 1:
                                discount_value = int(h_cell_value * 100)
                            else:
                                discount_value = int(h_cell_value)
                            print(f"找到Discount值: {discount_value}")
                            break
            
            print(f"成功读取Excel文件: {self.excel_file}")
            print(f"表头行: {self.header_row}")
            print(f"数据开始行: {self.start_row}")
            print(f"数据结束行: {actual_end_row}")
            print(f"共读取 {len(data)} 行数据")
            print(f"表头字段: {[h for h in headers if h is not None]}")
            
            return data, discount_value
        except ValueError as ve:
            # 表头验证失败，重新抛出异常
            print(f"Excel表头验证失败: {ve}")
            raise ve  # 重新抛出ValueError异常
        except Exception as e:
            print(f"读取Excel文件失败: {e}")
            return [], None
