#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建测试用的Excel文件
"""

import openpyxl


def create_test_excel():
    """创建测试用的Excel文件"""
    # 创建工作簿
    wb = openpyxl.Workbook()
    # 获取活动工作表
    sheet = wb.active
    
    # 设置表头
    headers = ['姓名', '年龄', '性别', '邮箱', '地址']
    for i, header in enumerate(headers, 1):
        sheet.cell(row=1, column=i, value=header)
    
    # 添加测试数据
    test_data = [
        ['张三', 25, '男', 'zhangsan@example.com', '北京市朝阳区'],
        ['李四', 30, '女', 'lisi@example.com', '上海市浦东新区'],
        ['王五', 28, '男', 'wangwu@example.com', '广州市天河区']
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            sheet.cell(row=row_idx, column=col_idx, value=value)
    
    # 保存文件
    wb.save('test_data.xlsx')
    print('测试Excel文件已创建: test_data.xlsx')


if __name__ == '__main__':
    create_test_excel()
